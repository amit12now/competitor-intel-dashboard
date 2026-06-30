# -*- coding: utf-8 -*-
"""Stage 7: write the formatted master Excel workbook with Atlas Copco branding."""
import pickle
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

with open("sheets_final.pkl", "rb") as f:
    SHEETS = pickle.load(f)

ATLAS_BLUE = "0099CC"
DARK_NAVY = "1B1F2A"
LIGHT_BLUE = "E6F5FA"
WHITE = "FFFFFF"

SHEET_ORDER = [
    "Settings", "Clean_Executive_Summary", "Clean_Competitor_Activity",
    "Clean_Channel_Activity", "Clean_Competitor_x_Channel", "Clean_Trend_May_vs_June",
    "Clean_Theme_Frequency", "Clean_Theme_by_Competitor", "Clean_Theme_by_Channel",
    "Clean_Theme_Trend_May_vs_June", "Clean_Product_Service_Signals",
    "Clean_Social_Conversation", "Clean_Conversation_Top_Themes", "Clean_Conversation_Sentiment",
    "Clean_Most_Mentioned_Brands", "Clean_PR_News_Events", "Clean_Hiring_Expansion",
    "Clean_Opportunities", "Clean_Raw_Data",
]

OUT_PATH = "../Competitor_Intelligence_Master.xlsx"

wb = Workbook()
wb.remove(wb.active)

header_font = Font(color=WHITE, bold=True, size=11, name="Calibri")
header_fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
title_font = Font(color=ATLAS_BLUE, bold=True, size=14, name="Calibri")
body_font = Font(size=10, name="Calibri")
thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

for sheet_name in SHEET_ORDER:
    sdf = SHEETS[sheet_name]
    ws = wb.create_sheet(sheet_name[:31])

    # Title row
    ws.cell(row=1, column=1, value=sheet_name.replace("Clean_", "").replace("_", " ")).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(sdf.columns), 1))
    ws.row_dimensions[1].height = 22

    start_row = 3
    for j, col in enumerate(sdf.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, rec) in enumerate(sdf.iterrows(), start=start_row + 1):
        for j, col in enumerate(sdf.columns, start=1):
            val = rec[col]
            if pd.isna(val):
                val = ""
            c = ws.cell(row=i, column=j, value=val)
            c.font = body_font
            c.border = thin_border
            c.alignment = Alignment(vertical="top", wrap_text=False)

    # Column widths (heuristic based on header + sample content length)
    for j, col in enumerate(sdf.columns, start=1):
        col_letter = get_column_letter(j)
        try:
            sample = sdf[col].astype(str).str.len()
            max_len = max(int(sample.head(200).max() or 0), len(str(col)))
        except Exception:
            max_len = len(str(col))
        width = min(max(max_len * 0.95 + 2, 12), 60)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Add a real Excel Table + autofilter for usability (skip if 0 rows or dup headers)
    n_rows = len(sdf)
    n_cols = max(len(sdf.columns), 1)
    if n_rows > 0 and len(set(sdf.columns)) == len(sdf.columns):
        last_col_letter = get_column_letter(n_cols)
        ref = f"A{start_row}:{last_col_letter}{start_row + n_rows}"
        tbl_name = "Tbl_" + "".join(ch for ch in sheet_name if ch.isalnum())[:25]
        table = Table(displayName=tbl_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(n_cols)}{start_row}"

    ws.sheet_view.showGridLines = False

wb.active = 0
wb.save(OUT_PATH)
print("Saved:", OUT_PATH)
